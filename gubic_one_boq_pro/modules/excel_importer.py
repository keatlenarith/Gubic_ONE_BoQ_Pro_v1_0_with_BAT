"""Excel workbook import and sheet/table discovery for BoQ workbooks."""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from config.settings import BOQ_KEYWORDS
from modules.utils import clean_text, normalize_header, to_number


@dataclass
class SheetProfile:
    name: str
    max_row: int
    max_column: int
    non_empty_cells: int
    detected_header_row: int | None
    sheet_type: str
    warnings: list[str]


def workbook_sheet_names(path: str | Path) -> list[str]:
    wb = load_workbook(path, read_only=True, data_only=True)
    names = list(wb.sheetnames)
    wb.close()
    return names


def inspect_workbook(path: str | Path) -> list[SheetProfile]:
    wb = load_workbook(path, read_only=True, data_only=True)
    profiles: list[SheetProfile] = []
    for ws in wb.worksheets:
        non_empty = 0
        first_rows: list[list[Any]] = []
        for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            row_values = list(row)
            non_empty += sum(1 for v in row_values if v not in (None, ""))
            if r_idx <= 60:
                first_rows.append(row_values)
        header = detect_header_row(first_rows)
        lower_name = ws.title.lower()
        sheet_type = "boq"
        if any(key in lower_name for key in ["sum", "summary"]):
            sheet_type = "summary"
        elif any(key in lower_name for key in ["raw", "material", "mat-list"]):
            sheet_type = "lookup"
        elif "area" in lower_name:
            sheet_type = "area"
        elif header is None:
            sheet_type = "unknown"
        warnings: list[str] = []
        if non_empty == 0:
            warnings.append("Empty sheet")
        if header is None and sheet_type == "boq":
            warnings.append("Could not confidently detect a BoQ header row")
        profiles.append(SheetProfile(ws.title, ws.max_row, ws.max_column, non_empty, header, sheet_type, warnings))
    wb.close()
    return profiles


def extract_metadata(path: str | Path) -> dict[str, Any]:
    """Read project metadata from the first meaningful cells in the workbook."""
    wb = load_workbook(path, read_only=True, data_only=True)
    meta: dict[str, Any] = {
        "project_name": "Imported BoQ Project",
        "location": "",
        "owner": "",
        "revision": "",
        "area_m2": None,
    }
    candidate_areas: list[tuple[int, float]] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 40), values_only=True):
            raw_cells = list(row)
            cells = [clean_text(v) for v in raw_cells if clean_text(v)]
            joined = " | ".join(cells)
            low = joined.lower()
            if "project" in low and meta["project_name"] == "Imported BoQ Project":
                for cell in cells:
                    if "project" in cell.lower() and ":" in cell:
                        meta["project_name"] = clean_text(cell.split(":", 1)[1]) or meta["project_name"]
                        break
            if "location" in low and not meta["location"]:
                for cell in cells:
                    if "location" in cell.lower() and ":" in cell:
                        meta["location"] = clean_text(cell.split(":", 1)[1])
                        break
            if "owner" in low or "ower" in low:
                for cell in cells:
                    if ("owner" in cell.lower() or "ower" in cell.lower()) and ":" in cell:
                        meta["owner"] = clean_text(cell.split(":", 1)[1])
                        break
            # Area detection: only accept values adjacent to an Area label or from Area sheet total rows.
            for idx, value in enumerate(raw_cells):
                text = clean_text(value).lower()
                if "area" in text and not any(skip in text for skip in ["using area", "area (", "surface"]):
                    priority = 0 if "area" in text and "=" in text else 1
                    for neighbor in raw_cells[idx + 1: idx + 4]:
                        num = to_number(neighbor)
                        if num and num >= 100:
                            candidate_areas.append((priority, float(num)))
                if ws.title.lower() == "area" and "total" in text:
                    for neighbor in raw_cells[idx + 1: idx + 4]:
                        num = to_number(neighbor)
                        if num and num >= 100:
                            candidate_areas.append((2, float(num)))
    if candidate_areas:
        plausible = [(p, a) for p, a in candidate_areas if 100 <= a <= 100000]
        priority = min(p for p, _ in plausible) if plausible else min(p for p, _ in candidate_areas)
        candidates = [a for p, a in (plausible or candidate_areas) if p == priority]
        # Within same priority, avoid very small mock-up sub-areas by taking the largest candidate below 20,000 m² where possible.
        building_scale = [a for a in candidates if 1000 <= a <= 20000]
        meta["area_m2"] = max(building_scale) if building_scale else max(candidates)
    wb.close()
    return meta


def detect_header_row(rows: list[list[Any]], max_scan_rows: int = 60) -> int | None:
    best_score = 0
    best_index: int | None = None
    for i, row in enumerate(rows[:max_scan_rows]):
        normalized = [normalize_header(v) for v in row if clean_text(v)]
        if not normalized:
            continue
        joined = " ".join(normalized)
        score = 0
        for keyword in BOQ_KEYWORDS:
            if keyword in joined:
                score += 1
        if "description" in joined or "descirption" in joined:
            score += 2
        if "unit" in joined and ("quantity" in joined or "qty" in joined):
            score += 2
        if score > best_score:
            best_score = score
            best_index = i
    return best_index if best_score >= 4 else None


def read_sheet_raw(path: str | Path, sheet_name: str) -> pd.DataFrame:
    return pd.read_excel(path, sheet_name=sheet_name, header=None, engine="openpyxl")


def read_all_raw(path: str | Path) -> dict[str, pd.DataFrame]:
    return pd.read_excel(path, sheet_name=None, header=None, engine="openpyxl")
